"""
ANVIQO CRITICAL SPARES INTELLIGENCE
===================================
Authoritative source: uploaded critical spares in pci.xlsx
This is SPARE-INSTRUMENT evidence, not spare PLC I/O.
Read-only. No PLC/SCADA control.
"""
from pathlib import Path
import re
import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "database" / "spares" / "critical_spares.xlsx"

SHEET_ALIASES = {
    "RTD": {"rtd", "thermocouple", "temperature", "te"},
    "LT": {"lt", "level", "level transmitter", "radar"},
    "PT": {"pt", "pressure", "pressure transmitter"},
    "FT": {"ft", "flow", "flow transmitter", "flow meter"},
    "MCV": {"mcv", "motorized control valve", "motorized valve"},
    "FSV&PCV": {"fsv", "pcv", "shut off valve", "pneumatic valve"},
    "SOV": {"sov", "solenoid", "solenoid valve"},
    "POS'R": {"positioner", "positioner", "posr"},
    "LOAD CELL": {"load cell", "loadcell", "encoder", "jam switch", "coal flow meter"},
    "Analyser": {"analyser", "analyzer", "gas analyser", "gas analyzer"},
    "AIR COMPRESSOR": {"air compressor", "compressor"},
    "Sheet1": {"general", "spares"},
}


def _norm(v):
    return re.sub(r"[^A-Z0-9]+", "", str(v or "").upper())


def _text(v):
    return str(v or "").strip()


def _parse_row(headers, values, sheet, row_no):
    d = {}
    for i, h in enumerate(headers):
        if h is not None and i < len(values):
            d[_text(h).strip().lower()] = values[i]
    # Common aliases
    tag = d.get('tag no') or d.get('tag') or d.get('tag no.')
    instrument = d.get('instrument') or d.get('short description') or d.get('description')
    spec = d.get('specification') or d.get('description')
    location = d.get('instalation site') or d.get('installation site') or d.get('location')
    qty = d.get('qty avbl')
    if qty is None:
        qty = d.get('qty')
    indent = d.get('spare to be indent')
    return {
        'sheet': sheet,
        'row': row_no,
        'tag': _text(tag),
        'instrument': _text(instrument),
        'description': _text(d.get('description') or d.get('item') or spec),
        'specification': _text(spec),
        'location': _text(location),
        'qty_available': qty,
        'spare_to_indent': indent,
        'raw': {k: v for k, v in d.items() if v not in (None, '')},
    }


def load_spares():
    if not XLSX.exists():
        return []
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        header_idx = 0
        for i, r in enumerate(all_rows[:6]):
            txt = ' '.join(_text(x).lower() for x in r if x is not None)
            if any(k in txt for k in ('tag no', 'description', 'short description', 'instrument', 'specification')):
                header_idx = i
                break
        headers = all_rows[header_idx]
        for rn, vals in enumerate(all_rows[header_idx+1:], header_idx+2):
            if not any(v not in (None, '') for v in vals):
                continue
            rec = _parse_row(headers, vals, sheet, rn)
            if rec['tag'] or rec['instrument'] or rec['description'] or rec['specification']:
                rows.append(rec)
    return rows


def _sheet_hint(q):
    """
    Deterministic instrument-family detection.

    IMPORTANT:
    Do not use generic substring matching for plant queries.
    Explicit instrument terminology gets priority.
    """

    ql = _text(q).lower()

    # Most specific phrases first.
    family_rules = [
        ("FSV&PCV", (
            "fsv",
            "pcv",
            "shut off valve",
            "shutoff valve",
            "pneumatic valve",
        )),
        ("AIR COMPRESSOR", (
            "air compressor",
            "compressor spare",
            "compressor spares",
        )),
        ("LOAD CELL", (
            "load cell",
            "loadcell",
            "coal flow meter",
            "jam switch",
        )),
        ("Analyser", (
            "analyser",
            "analyzer",
            "gas analyser",
            "gas analyzer",
        )),
        ("POS'R", (
            "positioner",
            "pos'r",
            "posr",
        )),
        ("MCV", (
            "mcv",
            "motorized control valve",
            "motorized valve",
        )),
        ("SOV", (
            "sov",
            "solenoid valve",
            "solenoid",
        )),
        ("RTD", (
            "rtd",
            "pt100",
            "temperature transmitter",
            "temperature sensor",
            "temperature element",
        )),
        ("LT", (
            "lt transmitter",
            "level transmitter",
            "level sensor",
            "radar level",
        )),
        ("PT", (
            "pt transmitter",
            "pressure transmitter",
            "pressure sensor",
            "pressure instrument",
        )),
        ("FT", (
            "ft transmitter",
            "flow transmitter",
            "flow sensor",
            "flow meter",
            "flowmeter",
        )),
    ]

    for sheet, phrases in family_rules:
        for phrase in phrases:
            if phrase in ql:
                return sheet

    # Exact tag-family recognition.
    import re

    for prefix, sheet in (
        ("MCV", "MCV"),
        ("PT", "PT"),
        ("FT", "FT"),
        ("LT", "LT"),
        ("RTD", "RTD"),
        ("TE", "RTD"),
        ("SOV", "SOV"),
        ("FSV", "FSV&PCV"),
        ("PCV", "FSV&PCV"),
    ):
        if re.search(r"\b" + re.escape(prefix) + r"[-_ ]?\d+[A-Z]?\b", ql):
            return sheet

    return None

def _qty_num(v):
    try:
        return float(v)
    except Exception:
        return None


def search_spares(query, limit=50):
    """
    Search critical spare instrument records.

    Family intent is enforced when a recognizable instrument family is
    present in the query. Availability intent is also enforced so that
    "available spares" does not return zero-stock records.
    """
    q = _text(query)
    ql = q.lower()
    qn = _norm(q)

    rows = load_spares()
    if not rows:
        return []

    sheet_hint = _sheet_hint(q)

    # ------------------------------------------------------------
    # EXACT SPARE TAG MATCH
    # ------------------------------------------------------------
    # Resolve explicit instrument tags before availability filtering.
    # This preserves zero-stock records such as:
    # PT-303 -> qty available 0, spare to indent 1.
    #
    # Therefore:
    # "Is PT-303 available as a spare?"
    # must return PT-303, not unrelated available spares.

    tag_pattern = re.compile(
        r"\b(MCV|PT|FT|LT|RTD|TE|SOV|FSV|PCV)[-_ ]?\d+[A-Z]?\b",
        re.IGNORECASE,
    )

    tag_match = tag_pattern.search(q)

    if tag_match:
        exact_tag_norm = _norm(tag_match.group(0))

        exact_rows = [
            r for r in rows
            if _norm(r.get("tag")) == exact_tag_norm
        ]

        if exact_rows:
            return exact_rows[:limit]

    # Explicit availability intent.
    available_only = any(x in ql for x in (
        "available spare",
        "available spares",
        "available",
        "in stock",
        "stock available",
        "currently available",
    ))

    unavailable_only = any(x in ql for x in (
        "not available",
        "unavailable",
        "out of stock",
        "to indent",
        "indent",
    ))

    # Remove conversational words for scoring.
    stop = {
        'DO','DI','AI','AO','THE','FOR','OF','A','AN','IS','ARE',
        'WE','HAVE','DOES','SPARE','SPARES','CRITICAL','AVAILABLE',
        'AVAILABILITY','SHOW','ME','TELL','WHAT','ABOUT','INSTRUMENT',
        'INSTRUMENTS','CURRENTLY','STOCK','IN','OUT','OF'
    }

    tokens = [
        _norm(x)
        for x in re.findall(r'[A-Za-z0-9_-]+', q)
        if _norm(x) not in stop
    ]

    # IMPORTANT:
    # If the user explicitly identifies an instrument family, restrict
    # the search to that source sheet. This prevents FT queries from
    # returning RTD/Analyser/other records.
    candidate_rows = rows

    if sheet_hint:
        candidate_rows = [
            r for r in rows
            if r['sheet'] == sheet_hint
        ]

    # Availability filtering.
    if available_only and not unavailable_only:
        filtered = []
        for r in candidate_rows:
            qty = _qty_num(r.get('qty_available'))
            if qty is not None and qty > 0:
                filtered.append(r)
        candidate_rows = filtered

    elif unavailable_only:
        filtered = []
        for r in candidate_rows:
            qty = _qty_num(r.get('qty_available'))
            if qty is None or qty <= 0:
                filtered.append(r)
        candidate_rows = filtered

    scored = []

    for r in candidate_rows:
        fields = ' '.join([
            r['tag'],
            r['instrument'],
            r['description'],
            r['specification'],
            r['location'],
            r['sheet']
        ])

        fn = _norm(fields)
        score = 0

        if sheet_hint and r['sheet'] == sheet_hint:
            score += 100

        if qn and qn in fn:
            score += 80

        for t in tokens:
            if t and t in fn:
                score += 15

        # Strong exact tag/family matching:
        # PT303 == PT-303, MCV204 == MCV-204.
        tag_norm = _norm(r['tag'])

        if tag_norm and any(
            _norm(t) == tag_norm for t in tokens
        ):
            score += 150

        # Family-specific boost.
        if sheet_hint:
            family_norm = _norm(sheet_hint)
            if family_norm and family_norm in fn:
                score += 50

        if score:
            scored.append((score, r))

    scored.sort(
        key=lambda x: (-x[0], x[1]['sheet'], x[1]['row'])
    )

    return [r for _, r in scored[:limit]]

def answer_spare_query(query):
    rows = search_spares(query)
    if not rows:
        return None
    lines = [f"ANVI found {len(rows)} verified critical spare instrument record(s). Source: critical spares in pci.xlsx."]
    for r in rows:
        qty = r['qty_available']
        indent = r['spare_to_indent']
        qty_text = f"Qty available: {qty}" if qty not in (None, '') else "Qty available: not recorded"
        ind_text = f"Spare to indent: {indent}" if indent not in (None, '') else "Spare to indent: not recorded"
        lines.append(
            f"{r['tag'] or r['instrument'] or 'UNNAMED'} — {r['instrument'] or r['description'] or 'Critical spare instrument'}; "
            f"Location: {r['location'] or 'not recorded'}; {qty_text}; {ind_text}; "
            f"Sheet: {r['sheet']}; Row: {r['row']}."
        )
        if r['specification']:
            lines.append(f"Specification: {r['specification']}")
    return {
        'answer': '\n'.join(lines),
        'domain': 'critical_spares',
        'evidence': 'critical spares in pci.xlsx',
        'count': len(rows),
        'records': rows,
        'read_only': True,
        'plc_write': False,
        'scada_control': False,
        'human_decision_required': True,
    }
